import hashlib
import io
import re
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import delete, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from server.utils.auth_middleware import get_admin_user, get_db
from src.storage.postgres.models_business import (
    CompliancePositionResponsibility,
    ComplianceProcessChecklist,
    ComplianceRiskLibrary,
    User,
)
from src.utils import logger

compliance = APIRouter(prefix="/compliance-risk", tags=["compliance-risk"])


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    text = _normalize_text(value)
    text = re.sub(r"[\s\u3000（）()【】\[\]:：\-—·/、,，。]", "", text)
    return text.lower()


def _split_items(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[\n\r；;|]+", value)
    clean = [p.strip() for p in parts if p and p.strip()]
    dedup: list[str] = []
    for item in clean:
        if item not in dedup:
            dedup.append(item)
    return dedup


def _safe_code(prefix: str, sheet: str, row_no: int, text: str) -> str:
    digest = hashlib.md5(f"{prefix}-{sheet}-{row_no}-{text}".encode("utf-8")).hexdigest()[:12]
    return f"{prefix}-{digest}"


def _cell(ws, row_no: int, col_no: int | None) -> str:
    if not col_no:
        return ""
    value = ws.cell(row=row_no, column=col_no).value
    if value is None:
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row_no <= merged.max_row and merged.min_col <= col_no <= merged.max_col:
                value = ws.cell(row=merged.min_row, column=merged.min_col).value
                break
    return _normalize_text(value)


def _find_columns(
    ws,
    column_candidates: dict[str, list[str]],
    required: set[str],
    max_scan_rows: int = 40,
) -> tuple[dict[str, int], int]:
    matched: dict[str, tuple[int, int]] = {}
    max_row = min(ws.max_row, max_scan_rows)

    for row_no in range(1, max_row + 1):
        for col_no in range(1, ws.max_column + 1):
            cell_norm = _normalize_header(ws.cell(row=row_no, column=col_no).value)
            if not cell_norm:
                continue
            for key, candidates in column_candidates.items():
                for candidate in candidates:
                    candidate_norm = _normalize_header(candidate)
                    if not candidate_norm:
                        continue
                    if candidate_norm == cell_norm or candidate_norm in cell_norm:
                        prev = matched.get(key)
                        # 优先保留更靠前的表头匹配，避免“其他/备注”等正文词命中后把 header_row 推到很后面。
                        if prev is None or row_no <= prev[0]:
                            matched[key] = (row_no, col_no)
                        break

    missing = [key for key in required if key not in matched]
    if missing:
        raise ValueError(f"表 {ws.title} 缺少必要列: {', '.join(missing)}")

    header_row = max(row for row, _ in matched.values())
    return {k: v[1] for k, v in matched.items()}, header_row


def _guess_category(*texts: str) -> str:
    joined = " ".join([t for t in texts if t]).lower()
    mapping = [
        ("法务 法律 合同 招标", "法律合规"),
        ("安全 作业 保密", "安全合规"),
        ("财务 预算 报销", "财务合规"),
        ("营销 电费", "营销合规"),
        ("廉洁 纪检", "廉洁合规"),
        ("信息 数据 网络", "信息合规"),
        ("工程 建设", "工程合规"),
    ]
    for words, category in mapping:
        if any(word in joined for word in words.split()):
            return category
    return "综合合规"


def _parse_risk_library(file_name: str, workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        columns, header_row = _find_columns(
            ws,
            column_candidates={
                "business_lv1": ["一级业务"],
                "business_lv2": ["二级业务"],
                "title": ["合规风险名称"],
                "code": ["风险行为编号"],
                "desc": ["风险行为描述"],
                "consequence": ["责任或后果"],
                "level": ["风险等级"],
                "bottom_line": ["底线"],
                "basis_policy": ["国家政策"],
                "basis_law": ["法律法规"],
                "basis_regulation": ["监管规定"],
                "basis_industry": ["行业准则"],
                "basis_internal": ["规章制度"],
                "basis_other": ["其他"],
                "obligation": ["合规义务"],
                "measures": ["风险控制措施"],
                "department": ["归口部门"],
                "cooperate_department": ["配合部门"],
                "remark": ["备注"],
            },
            required={"title", "code", "desc"},
        )

        business_lv1 = ""
        business_lv2 = ""
        department = ""
        for row_no in range(header_row + 1, ws.max_row + 1):
            title = _cell(ws, row_no, columns.get("title"))
            code = _cell(ws, row_no, columns.get("code"))
            desc_text = _cell(ws, row_no, columns.get("desc"))
            if not any([title, code, desc_text]):
                continue
            if title in {"合规风险名称", "风险行为描述"}:
                continue

            code = (code.splitlines()[0].strip() if code else "") or _safe_code("FX", ws.title, row_no, f"{title}-{desc_text}")
            raw_level = _cell(ws, row_no, columns.get("level"))
            if "高" in raw_level:
                level = "高风险"
            elif "中" in raw_level:
                level = "中风险"
            elif "低" in raw_level:
                level = "低风险"
            else:
                level = raw_level[:20]
            business_lv1 = _cell(ws, row_no, columns.get("business_lv1")) or business_lv1
            business_lv2 = _cell(ws, row_no, columns.get("business_lv2")) or business_lv2
            department = _cell(ws, row_no, columns.get("department")) or department
            basis_parts = [
                _cell(ws, row_no, columns.get("basis_policy")),
                _cell(ws, row_no, columns.get("basis_law")),
                _cell(ws, row_no, columns.get("basis_regulation")),
                _cell(ws, row_no, columns.get("basis_industry")),
                _cell(ws, row_no, columns.get("basis_internal")),
                _cell(ws, row_no, columns.get("basis_other")),
            ]
            basis = "\n".join([item for item in basis_parts if item])
            measures = _split_items(_cell(ws, row_no, columns.get("measures")))
            chips = [item for item in [business_lv1, business_lv2, department] if item]

            rows.append(
                {
                    "code": code,
                    "title": title or code,
                    "desc": desc_text,
                    "level": level,
                    "category": _guess_category(title, desc_text, business_lv1, business_lv2),
                    "department": department,
                    "cooperate_department": _cell(ws, row_no, columns.get("cooperate_department")),
                    "owner": "",
                    "business_lv1": business_lv1,
                    "business_lv2": business_lv2,
                    "consequence": _cell(ws, row_no, columns.get("consequence")),
                    "bottom_line": _cell(ws, row_no, columns.get("bottom_line")),
                    "basis": basis,
                    "obligation": _cell(ws, row_no, columns.get("obligation")),
                    "measures": measures,
                    "chips": chips,
                    "remark": _cell(ws, row_no, columns.get("remark")),
                    "source_sheet": ws.title,
                    "source_file": file_name,
                }
            )
    return rows


def _parse_process_checklist(file_name: str, workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        columns, header_row = _find_columns(
            ws,
            column_candidates={
                "level1_process": ["一级流程"],
                "level2_process": ["二级流程"],
                "level3_process": ["三级流程"],
                "code": ["末级流程编号"],
                "title": ["末级流程名称"],
                "risk_desc": ["合规重要环节", "主要风险名称"],
                "compliance_points": ["合规审查内容", "关键控制点"],
                "source_external": ["法律法规国家政策", "法律法规、国家政策"],
                "source_internal": ["内部制度"],
                "risk_points": ["合规风险点"],
                "measures": ["监督评价要点"],
                "department": ["合规审查责任部门"],
            },
            required={"code", "title"},
        )

        level1_process = ""
        level2_process = ""
        level3_process = ""
        department = ""
        for row_no in range(header_row + 1, ws.max_row + 1):
            code = _cell(ws, row_no, columns.get("code"))
            title = _cell(ws, row_no, columns.get("title"))
            if not code and not title:
                continue
            if code == "末级流程编号":
                continue

            code = (code.splitlines()[0].strip() if code else "") or _safe_code("LC", ws.title, row_no, title)
            level1_process = _cell(ws, row_no, columns.get("level1_process")) or level1_process
            level2_process = _cell(ws, row_no, columns.get("level2_process")) or level2_process
            level3_process = _cell(ws, row_no, columns.get("level3_process")) or level3_process
            department = _cell(ws, row_no, columns.get("department")) or department
            risk_desc = _cell(ws, row_no, columns.get("risk_desc"))
            risk_points = _cell(ws, row_no, columns.get("risk_points"))
            source_basis = "\n".join(
                [
                    item
                    for item in [
                        _cell(ws, row_no, columns.get("source_external")),
                        _cell(ws, row_no, columns.get("source_internal")),
                    ]
                    if item
                ]
            )
            compliance_points_text = _cell(ws, row_no, columns.get("compliance_points"))
            compliance_points = _split_items(compliance_points_text)

            rows.append(
                {
                    "code": code,
                    "title": title or code,
                    "department": department,
                    "owner": "",
                    "level1_process": level1_process,
                    "level2_process": level2_process,
                    "level3_process": level3_process,
                    "risk_desc": risk_desc or risk_points,
                    "compliance_points": compliance_points,
                    "source_basis": source_basis,
                    "risk_points": risk_points,
                    "measures": _cell(ws, row_no, columns.get("measures")),
                    "source_sheet": ws.title,
                    "source_file": file_name,
                }
            )
    return rows


def _parse_position_responsibility(file_name: str, workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        columns, header_row = _find_columns(
            ws,
            column_candidates={
                "department": ["部门名称"],
                "title": ["岗位名称"],
                "responsibilities": ["风险内控合规职责"],
                "source_external": ["法律法规国家政策", "法律法规、国家政策"],
                "source_internal": ["内部制度"],
                "compliance_points": ["合规底线清单"],
                "bottom_line_penalty": ["底线标准与处罚"],
                "related_risks": ["制度依据"],
            },
            required={"department", "title"},
        )

        department = ""
        for row_no in range(header_row + 1, ws.max_row + 1):
            department = _cell(ws, row_no, columns.get("department")) or department
            title = _cell(ws, row_no, columns.get("title"))
            if not department and not title:
                continue
            if title == "岗位名称":
                continue

            code = _safe_code("GW", ws.title, row_no, f"{department}-{title}")
            source_basis = "\n".join(
                [
                    item
                    for item in [
                        _cell(ws, row_no, columns.get("source_external")),
                        _cell(ws, row_no, columns.get("source_internal")),
                    ]
                    if item
                ]
            )

            rows.append(
                {
                    "code": code,
                    "title": title or code,
                    "department": department,
                    "compliance_type": _guess_category(title, department),
                    "level": "",
                    "responsibilities": _split_items(_cell(ws, row_no, columns.get("responsibilities"))),
                    "compliance_points": _split_items(_cell(ws, row_no, columns.get("compliance_points"))),
                    "bottom_line_penalty": _cell(ws, row_no, columns.get("bottom_line_penalty")),
                    "source_basis": source_basis,
                    "related_risks": _split_items(_cell(ws, row_no, columns.get("related_risks"))),
                    "source_sheet": ws.title,
                    "source_file": file_name,
                }
            )
    return rows


class ImportResult(BaseModel):
    data_type: str
    imported: int
    updated: int
    total: int


@compliance.post("/import/excel")
async def import_compliance_excel(
    data_type: str = Form(...),
    replace_existing: bool = Form(True),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")

    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        logger.exception("读取 Excel 失败")
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {exc}") from exc

    data_type = data_type.strip().lower()
    parser_map = {
        "risk-library": _parse_risk_library,
        "process-checklist": _parse_process_checklist,
        "position-responsibility": _parse_position_responsibility,
    }
    model_map = {
        "risk-library": ComplianceRiskLibrary,
        "process-checklist": ComplianceProcessChecklist,
        "position-responsibility": CompliancePositionResponsibility,
    }

    parser = parser_map.get(data_type)
    model_cls = model_map.get(data_type)
    if not parser or not model_cls:
        raise HTTPException(status_code=400, detail="不支持的数据类型")

    try:
        parsed_rows = parser(filename, workbook)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Excel 解析业务数据失败")
        raise HTTPException(status_code=500, detail=f"导入解析失败: {exc}") from exc

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="未解析到有效数据，请检查表头与内容")

    dedup_rows_map: dict[str, dict[str, Any]] = {}
    for row in parsed_rows:
        dedup_rows_map[row["code"]] = row
    dedup_rows = list(dedup_rows_map.values())

    imported = 0
    updated = 0

    if replace_existing:
        await db.execute(delete(model_cls))

    for row in dedup_rows:
        row["updated_by"] = str(current_user.id)
        if replace_existing:
            row["created_by"] = str(current_user.id)
            db.add(model_cls(**row))
            imported += 1
            continue

        existing = await db.execute(select(model_cls).where(model_cls.code == row["code"]))
        entity = existing.scalar_one_or_none()
        if entity is None:
            row["created_by"] = str(current_user.id)
            db.add(model_cls(**row))
            imported += 1
            continue

        for key, value in row.items():
            setattr(entity, key, value)
        updated += 1

    await db.flush()
    return ImportResult(data_type=data_type, imported=imported, updated=updated, total=len(dedup_rows)).model_dump()


@compliance.get("/risk-library")
async def list_risk_library(
    keyword: str | None = None,
    level: str | None = None,
    department: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    stmt = select(ComplianceRiskLibrary)
    if keyword:
        kw = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                ComplianceRiskLibrary.code.ilike(kw),
                ComplianceRiskLibrary.title.ilike(kw),
                ComplianceRiskLibrary.desc.ilike(kw),
                ComplianceRiskLibrary.department.ilike(kw),
            )
        )
    if level:
        stmt = stmt.where(ComplianceRiskLibrary.level == level)
    if department:
        stmt = stmt.where(ComplianceRiskLibrary.department == department)
    result = await db.execute(stmt.order_by(desc(ComplianceRiskLibrary.updated_at), desc(ComplianceRiskLibrary.id)))
    data = [row.to_dict() for row in result.scalars().all()]
    return {"data": data}


@compliance.get("/risk-library/{record_id}")
async def get_risk_library_detail(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    result = await db.execute(select(ComplianceRiskLibrary).where(ComplianceRiskLibrary.id == record_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="记录不存在")
    return {"data": row.to_dict()}


@compliance.get("/process-checklist")
async def list_process_checklist(
    keyword: str | None = None,
    department: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    stmt = select(ComplianceProcessChecklist)
    if keyword:
        kw = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                ComplianceProcessChecklist.code.ilike(kw),
                ComplianceProcessChecklist.title.ilike(kw),
                ComplianceProcessChecklist.risk_desc.ilike(kw),
                ComplianceProcessChecklist.department.ilike(kw),
            )
        )
    if department:
        stmt = stmt.where(ComplianceProcessChecklist.department == department)
    result = await db.execute(
        stmt.order_by(desc(ComplianceProcessChecklist.updated_at), desc(ComplianceProcessChecklist.id))
    )
    data = [row.to_dict() for row in result.scalars().all()]
    return {"data": data}


@compliance.get("/process-checklist/{record_id}")
async def get_process_checklist_detail(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    result = await db.execute(select(ComplianceProcessChecklist).where(ComplianceProcessChecklist.id == record_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="记录不存在")
    return {"data": row.to_dict()}


@compliance.get("/position-responsibility")
async def list_position_responsibility(
    keyword: str | None = None,
    department: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    stmt = select(CompliancePositionResponsibility)
    if keyword:
        kw = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                CompliancePositionResponsibility.code.ilike(kw),
                CompliancePositionResponsibility.title.ilike(kw),
                CompliancePositionResponsibility.department.ilike(kw),
                CompliancePositionResponsibility.compliance_type.ilike(kw),
            )
        )
    if department:
        stmt = stmt.where(CompliancePositionResponsibility.department == department)
    result = await db.execute(
        stmt.order_by(desc(CompliancePositionResponsibility.updated_at), desc(CompliancePositionResponsibility.id))
    )
    data = [row.to_dict() for row in result.scalars().all()]
    return {"data": data}


@compliance.get("/position-responsibility/{record_id}")
async def get_position_responsibility_detail(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    result = await db.execute(select(CompliancePositionResponsibility).where(CompliancePositionResponsibility.id == record_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="记录不存在")
    return {"data": row.to_dict()}


@compliance.get("/summary")
async def get_compliance_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    risk_count = await db.scalar(select(func.count()).select_from(ComplianceRiskLibrary))
    process_count = await db.scalar(select(func.count()).select_from(ComplianceProcessChecklist))
    position_count = await db.scalar(select(func.count()).select_from(CompliancePositionResponsibility))

    recent_risk = (
        await db.execute(select(ComplianceRiskLibrary).order_by(desc(ComplianceRiskLibrary.updated_at)).limit(5))
    ).scalars().all()
    recent_process = (
        await db.execute(select(ComplianceProcessChecklist).order_by(desc(ComplianceProcessChecklist.updated_at)).limit(5))
    ).scalars().all()
    recent_position = (
        await db.execute(
            select(CompliancePositionResponsibility).order_by(desc(CompliancePositionResponsibility.updated_at)).limit(5)
        )
    ).scalars().all()

    updates = []
    updates.extend(
        [
            {
                "id": f"risk-{item.id}",
                "type": "风险库",
                "title": item.title,
                "role": item.business_lv2 or item.business_lv1 or "",
                "department": item.department or "",
                "updated_at": item.updated_at,
            }
            for item in recent_risk
        ]
    )
    updates.extend(
        [
            {
                "id": f"proc-{item.id}",
                "type": "流程清单",
                "title": item.title,
                "role": item.level2_process or item.level1_process or "",
                "department": item.department or "",
                "updated_at": item.updated_at,
            }
            for item in recent_process
        ]
    )
    updates.extend(
        [
            {
                "id": f"pos-{item.id}",
                "type": "岗位清单",
                "title": item.title,
                "role": item.compliance_type or "",
                "department": item.department or "",
                "updated_at": item.updated_at,
            }
            for item in recent_position
        ]
    )

    updates = sorted(updates, key=lambda x: x["updated_at"] or 0, reverse=True)[:10]
    for item in updates:
        item["updated_at"] = item["updated_at"].isoformat() if item["updated_at"] else ""

    return {
        "data": {
            "counts": {
                "risk_library": risk_count or 0,
                "process_checklist": process_count or 0,
                "position_responsibility": position_count or 0,
            },
            "recent_updates": updates,
        }
    }
