from openpyxl import Workbook

from server.routers.compliance_router import _cell, _find_columns, _parse_process_checklist


def test_find_columns_keeps_header_row_when_body_contains_header_like_words() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试sheet"

    ws["A2"] = "一级业务"
    ws["B2"] = "二级业务"
    ws["C2"] = "合规风险名称"
    ws["D2"] = "风险行为编号"
    ws["E2"] = "风险行为描述"
    ws["F2"] = "备注"
    ws["G2"] = "其他"

    # 正文出现“备注/其他”，不应把表头定位错到这里。
    ws["F20"] = "备注"
    ws["G21"] = "其他"

    columns, header_row = _find_columns(
        ws,
        column_candidates={
            "business_lv1": ["一级业务"],
            "business_lv2": ["二级业务"],
            "title": ["合规风险名称"],
            "code": ["风险行为编号"],
            "desc": ["风险行为描述"],
            "remark": ["备注"],
            "basis_other": ["其他"],
        },
        required={"title", "code", "desc"},
        max_scan_rows=30,
    )

    assert header_row == 2
    assert columns["code"] == 4
    assert columns["desc"] == 5


def test_cell_reads_from_merged_anchor() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A3"] = "后勤保障部"
    ws.merge_cells("A3:A8")

    assert _cell(ws, 6, 1) == "后勤保障部"


def test_parse_process_checklist_fills_merged_process_levels() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "综合服务中心"

    ws["A2"] = "一级流程"
    ws["B2"] = "二级流程"
    ws["C2"] = "三级流程"
    ws["D2"] = "末级流程编号"
    ws["E2"] = "末级流程名称"
    ws["F2"] = "合规重要环节"
    ws["G2"] = "合规审查内容"
    ws["H2"] = "法律法规、国家政策"
    ws["I2"] = "内部制度"
    ws["J2"] = "合规风险点"
    ws["K2"] = "监督评价要点"
    ws["L2"] = "合规审查责任部门"

    ws["A3"] = "后勤保障部"
    ws["B3"] = "食堂餐饮管理"
    ws["C3"] = "食堂餐饮管理"
    ws.merge_cells("A3:A5")
    ws.merge_cells("B3:B5")
    ws.merge_cells("C3:C5")

    ws["D3"] = "HQ-08-01-07"
    ws["E3"] = "食品卫生安全管理"
    ws["L3"] = "综合服务中心"

    rows = _parse_process_checklist("x.xlsx", wb)
    assert len(rows) == 1
    row = rows[0]
    assert row["code"] == "HQ-08-01-07"
    assert row["level1_process"] == "后勤保障部"
    assert row["level2_process"] == "食堂餐饮管理"
    assert row["level3_process"] == "食堂餐饮管理"
