from pathlib import Path

from openpyxl import Workbook

from src.services.jingzhou_compliance_seed_service import JingzhouComplianceSeedService


def test_extract_process_records_fill_merged_cells_from_anchor_row(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "综合服务中心"

    # 模拟合并单元格锚点在第3行，解析从第4行开始。
    ws["A3"] = "后勤保障部"
    ws["B3"] = "食堂餐饮管理"
    ws["C3"] = "食堂餐饮管理"
    ws.merge_cells("A3:A10")
    ws.merge_cells("B3:B10")
    ws.merge_cells("C3:C10")

    ws["D4"] = "HQ-08-01-07"
    ws["E4"] = "食品卫生安全管理"
    ws["F4"] = "R-HQ-07-04"
    ws["G4"] = "无"
    ws["H4"] = "外部依据"
    ws["I4"] = "内部依据"
    ws["J4"] = "风险点"

    records = JingzhouComplianceSeedService._extract_process_records(ws, ws.title)
    assert len(records) == 1
    record = records[0]
    assert "- 一级流程: 后勤保障部" in record
    assert "- 二级流程: 食堂餐饮管理" in record
    assert "- 三级流程: 食堂餐饮管理" in record
    assert "- 末级流程编号: HQ-08-01-07" in record


def test_convert_workbook_keeps_empty_sheet_md(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "空Sheet"
    ws["A1"] = "仅用于触发输出"

    source_file = tmp_path / "empty.xlsx"
    wb.save(source_file)

    output_dir = tmp_path / "out"
    produced = JingzhouComplianceSeedService._convert_workbook(source_file, output_dir, "process")

    assert len(produced) == 1
    out_file = produced[0]
    assert out_file.name == "空Sheet.md"
    content = out_file.read_text(encoding="utf-8")
    assert "# 空Sheet" in content
    assert "- 记录数: 0" in content
    assert "本 sheet 未抽取到可入库记录" in content
