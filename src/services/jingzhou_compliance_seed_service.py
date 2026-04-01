from __future__ import annotations

import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src import config
from src.knowledge import knowledge_base
from src.knowledge.base import FileStatus
from src.repositories.knowledge_base_repository import KnowledgeBaseRepository
from src.services.task_service import TaskContext, tasker
from src.storage.postgres.manager import pg_manager
from src.utils import logger


@dataclass
class ComplianceSeedResult:
    kb_name: str
    db_id: str
    total_files: int
    imported_files: int
    skipped_files: int
    message: str


class JingzhouComplianceSeedService:
    """荆州一库两清单种子导入服务（支持启动自动导入 + 手动补跑）"""

    SOURCE_DIR_ENV = "YUXI_JINGZHOU_COMPLIANCE_SOURCE_DIR"
    AUTO_SEED_ENV = "YUXI_JINGZHOU_COMPLIANCE_AUTO_SEED"
    FORCE_ENV = "YUXI_JINGZHOU_COMPLIANCE_FORCE"
    CHUNK_SIZE_ENV = "YUXI_JINGZHOU_COMPLIANCE_CHUNK_SIZE"
    CHUNK_OVERLAP_ENV = "YUXI_JINGZHOU_COMPLIANCE_CHUNK_OVERLAP"
    INDEX_CONCURRENCY_ENV = "YUXI_JINGZHOU_COMPLIANCE_INDEX_CONCURRENCY"
    SEARCH_MODE_ENV = "YUXI_JINGZHOU_COMPLIANCE_SEARCH_MODE"
    TOP_K_ENV = "YUXI_JINGZHOU_COMPLIANCE_TOP_K"
    KEYWORD_TOP_K_ENV = "YUXI_JINGZHOU_COMPLIANCE_KEYWORD_TOP_K"

    KB_DEFS = [
        {
            "kb_name": "合规风险库",
            "description": "荆州供电公司-合规风险库（由系统种子脚本自动导入）",
            "source_file": "1.国网荆州供电公司合规风险库.xlsx",
            "kind": "risk",
        },
        {
            "kb_name": "岗位义务清单",
            "description": "荆州供电公司-岗位义务清单（由系统种子脚本自动导入）",
            "source_file": "2.国网荆州供电公司重要岗位履责合规义务清单.xlsx",
            "kind": "duty",
        },
        {
            "kb_name": "流程管控清单",
            "description": "荆州供电公司-流程管控清单（由系统种子脚本自动导入）",
            "source_file": "3.国网荆州供电公司业务流程管控合规管理清单.xlsx",
            "kind": "process",
        },
    ]

    @classmethod
    def _source_dir_candidates(cls) -> list[Path]:
        configured = os.getenv(cls.SOURCE_DIR_ENV)
        candidates = []
        if configured:
            candidates.append(Path(configured).expanduser())
        candidates.append(Path(config.save_dir) / "jingzhou_compliance_seed" / "source")
        candidates.append(Path.cwd() / "data" / "seed" / "jingzhou_compliance" / "source")
        candidates.append(
            Path("/Users/wangying/Documents/外包/9-罗总-惠州电力局人工智能/2025年国网荆州供电公司“一库两清单”")
        )
        deduped: list[Path] = []
        seen: set[str] = set()
        for c in candidates:
            key = str(c)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(c)
        return deduped

    @classmethod
    def _resolve_source_dir(cls) -> Path | None:
        for path in cls._source_dir_candidates():
            if path.exists() and path.is_dir():
                return path
        return None

    @classmethod
    def _generated_root(cls) -> Path:
        return Path(config.save_dir) / "jingzhou_compliance_seed" / "generated"

    @staticmethod
    def _sanitize_name(name: str) -> str:
        text = re.sub(r"[\\/:*?\"<>|]+", "_", name.strip())
        text = re.sub(r"\s+", "_", text)
        return text[:128] or "unknown"

    @classmethod
    def _resolve_chunk_params(cls) -> dict:
        chunk_size = int(os.getenv(cls.CHUNK_SIZE_ENV, "2200"))
        chunk_overlap = int(os.getenv(cls.CHUNK_OVERLAP_ENV, "200"))
        index_concurrency = int(os.getenv(cls.INDEX_CONCURRENCY_ENV, "4"))
        return {
            "chunk_size": max(256, chunk_size),
            "chunk_overlap": max(0, min(chunk_overlap, max(0, chunk_size - 1))),
            "index_concurrency": max(1, index_concurrency),
            "qa_separator": "",
            "content_type": "file",
            "auto_index": False,
        }

    @classmethod
    async def enqueue_startup_seed(cls, operator_id: int | str = 1, department_id: int | None = 1) -> None:
        if os.getenv(cls.AUTO_SEED_ENV, "true").strip().lower() not in {"1", "true", "yes", "on"}:
            logger.info("Jingzhou compliance startup seed disabled by {}", cls.AUTO_SEED_ENV)
            return

        source_dir = cls._resolve_source_dir()
        if source_dir is None:
            logger.info("Jingzhou compliance startup seed skipped: source dir not found")
            return

        async def run_seed(context: TaskContext):
            await context.set_progress(5.0, "启动荆州一库两清单自动导入")
            results = await cls.seed_all(
                operator_id=str(operator_id),
                department_id=department_id,
                force=os.getenv(cls.FORCE_ENV, "false").strip().lower() in {"1", "true", "yes", "on"},
                source_dir=source_dir,
                context=context,
            )
            await context.set_progress(100.0, "荆州一库两清单自动导入完成")
            return [r.__dict__ for r in results]

        await tasker.enqueue(
            name="启动导入-荆州一库两清单",
            task_type="knowledge_seed_jingzhou_compliance",
            payload={"source_dir": str(source_dir)},
            coroutine=run_seed,
        )
        logger.info("Jingzhou compliance startup seed task enqueued, source_dir={}", source_dir)

    @classmethod
    async def seed_all(
        cls,
        operator_id: str,
        department_id: int | None,
        force: bool = False,
        source_dir: Path | None = None,
        context: TaskContext | None = None,
    ) -> list[ComplianceSeedResult]:
        if not pg_manager._initialized:
            pg_manager.initialize()
        await pg_manager.create_business_tables()
        await pg_manager.ensure_knowledge_schema()
        await knowledge_base.initialize()
        src_dir = source_dir or cls._resolve_source_dir()
        if src_dir is None:
            raise FileNotFoundError("未找到荆州一库两清单源目录，请检查 YUXI_JINGZHOU_COMPLIANCE_SOURCE_DIR")

        logger.info("Jingzhou compliance seed source_dir={}", src_dir)
        generated_root = cls._generated_root()
        generated_root.mkdir(parents=True, exist_ok=True)

        chunk_params = cls._resolve_chunk_params()
        results: list[ComplianceSeedResult] = []

        total_defs = len(cls.KB_DEFS)
        for idx, kb_def in enumerate(cls.KB_DEFS, 1):
            kb_name = kb_def["kb_name"]
            source_file = src_dir / kb_def["source_file"]
            if not source_file.exists():
                logger.warning("Seed source file missing for {}: {}", kb_name, source_file)
                continue

            if context:
                await context.set_progress(10.0 + (idx - 1) / total_defs * 80.0, f"处理知识库：{kb_name}")

            db_id = await cls._ensure_kb(
                kb_name=kb_name,
                description=kb_def["description"],
                operator_id=operator_id,
                department_id=department_id,
            )
            await cls._ensure_query_params(db_id)
            kb_output_dir = generated_root / cls._sanitize_name(kb_name)
            converted_files = cls._convert_workbook(source_file, kb_output_dir, kb_def["kind"])

            imported_count, skipped_count = await cls._import_generated_files(
                db_id=db_id,
                files=converted_files,
                operator_id=operator_id,
                chunk_params=chunk_params,
                force=force,
            )

            results.append(
                ComplianceSeedResult(
                    kb_name=kb_name,
                    db_id=db_id,
                    total_files=len(converted_files),
                    imported_files=imported_count,
                    skipped_files=skipped_count,
                    message=f"导入完成 imported={imported_count}, skipped={skipped_count}",
                )
            )

        return results

    @classmethod
    async def _ensure_kb(
        cls,
        kb_name: str,
        description: str,
        operator_id: str,
        department_id: int | None,
    ) -> str:
        repo = KnowledgeBaseRepository()
        rows = await repo.get_all()
        for row in rows:
            if (row.name or "").strip() == kb_name:
                return row.db_id

        embed_info = config.embed_model_names[config.embed_model]
        embed_info_dict = embed_info.model_dump() if hasattr(embed_info, "model_dump") else embed_info.dict()
        llm_info = {
            "provider": "",
            "model_name": (config.default_model or "").split("/", 1)[-1],
            "model_spec": config.default_model or "",
        }
        share_config = {
            "is_shared": True,
            "accessible_departments": [department_id] if department_id else [],
        }
        created = await knowledge_base.create_database(
            database_name=kb_name,
            description=description,
            kb_type="milvus",
            embed_info=embed_info_dict,
            llm_info=llm_info,
            share_config=share_config,
            auto_generate_questions=False,
        )
        logger.info("Created jingzhou compliance KB: {} -> {}", kb_name, created["db_id"])
        return created["db_id"]

    @classmethod
    async def _ensure_query_params(cls, db_id: str) -> None:
        search_mode = os.getenv(cls.SEARCH_MODE_ENV, "hybrid").strip() or "hybrid"
        top_k = max(1, int(os.getenv(cls.TOP_K_ENV, "8")))
        keyword_top_k = max(1, int(os.getenv(cls.KEYWORD_TOP_K_ENV, "30")))
        params = {
            "search_mode": search_mode,
            "top_k": top_k,
            "keyword_top_k": keyword_top_k,
        }

        kb_instance = await knowledge_base._get_kb_for_database(db_id)
        if db_id not in kb_instance.databases_meta:
            await kb_instance._load_metadata()
        if db_id not in kb_instance.databases_meta:
            logger.warning("Skip setting query params: db_id not found in metadata after reload: {}", db_id)
            return
        async with knowledge_base._metadata_lock:
            if kb_instance.databases_meta[db_id].get("query_params") is None:
                kb_instance.databases_meta[db_id]["query_params"] = {}
            options = kb_instance.databases_meta[db_id]["query_params"].setdefault("options", {})
            options.update(params)
            await kb_instance._save_metadata()

    @classmethod
    def _convert_workbook(cls, source_file: Path, output_dir: Path, kind: str) -> list[Path]:
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(source_file, data_only=True)
        produced: list[Path] = []

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            records: list[str] = []
            if kind == "risk":
                records = cls._extract_risk_records(ws, sheet_name)
            elif kind == "duty":
                records = cls._extract_duty_records(ws, sheet_name)
            elif kind == "process":
                records = cls._extract_process_records(ws, sheet_name)

            out_file = output_dir / f"{cls._sanitize_name(sheet_name)}.md"
            content = [
                f"# {sheet_name}",
                f"- 来源文件: {source_file.name}",
                f"- 记录数: {len(records)}",
                "",
            ]
            if records:
                content.extend(records)
            else:
                # 不再静默跳过空 sheet，避免“sheet 丢失”错觉，便于排查源数据与映射问题
                content.extend(
                    [
                        "> 本 sheet 未抽取到可入库记录，请检查表头行、编码列与合并单元格映射。",
                        "",
                    ]
                )
            out_file.write_text("\n".join(content), encoding="utf-8")
            produced.append(out_file)

        return produced

    @staticmethod
    def _cell(ws, row: int, col: int) -> str:
        value = ws.cell(row=row, column=col).value
        if value is None:
            # 对合并单元格做回填：若命中 merged range，读取左上角单元格值
            for merged in ws.merged_cells.ranges:
                if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                    value = ws.cell(row=merged.min_row, column=merged.min_col).value
                    break
        if value is None:
            return ""
        return str(value).strip()

    @classmethod
    def _extract_risk_records(cls, ws, sheet_name: str) -> list[str]:
        records: list[str] = []
        biz_l1 = ""
        biz_l2 = ""
        for row in range(6, ws.max_row + 1):
            biz_l1 = cls._cell(ws, row, 1) or biz_l1
            biz_l2 = cls._cell(ws, row, 2) or biz_l2
            risk_name = cls._cell(ws, row, 3)
            risk_code = cls._cell(ws, row, 4)
            risk_desc = cls._cell(ws, row, 5)
            consequence = cls._cell(ws, row, 6)
            risk_level = cls._cell(ws, row, 7)
            bottom_line = cls._cell(ws, row, 8)
            basis_policy = cls._cell(ws, row, 9)
            basis_law = cls._cell(ws, row, 10)
            basis_reg = cls._cell(ws, row, 11)
            basis_industry = cls._cell(ws, row, 12)
            basis_internal = cls._cell(ws, row, 13)
            basis_other = cls._cell(ws, row, 14)

            if not risk_code:
                continue

            lines = [
                "## 标准检索单元",
                f"- 数据类型: 合规风险",
                f"- 部门: {sheet_name}",
                f"- 一级业务: {biz_l1}",
                f"- 二级业务: {biz_l2}",
                f"- 风险编号: {risk_code}",
                f"- 风险名称: {risk_name}",
                f"- 风险行为描述: {risk_desc}",
                f"- 责任或后果: {consequence}",
                f"- 风险等级: {risk_level}",
                f"- 底线标记: {bottom_line}",
                f"- 合规依据(国家政策): {basis_policy}",
                f"- 合规依据(法律法规): {basis_law}",
                f"- 合规依据(监管规定): {basis_reg}",
                f"- 合规依据(行业准则): {basis_industry}",
                f"- 合规依据(内部制度): {basis_internal}",
                f"- 合规依据(其他): {basis_other}",
                "---",
                "",
            ]
            records.append("\n".join(lines))
        return records

    @classmethod
    def _extract_duty_records(cls, ws, sheet_name: str) -> list[str]:
        records: list[str] = []
        dept_name = ""
        for row in range(5, ws.max_row + 1):
            dept_name = cls._cell(ws, row, 1) or dept_name
            role_name = cls._cell(ws, row, 2)
            duty = cls._cell(ws, row, 3)
            source_external = cls._cell(ws, row, 4)
            source_internal = cls._cell(ws, row, 5)
            bottom_line = cls._cell(ws, row, 6)
            penalty = cls._cell(ws, row, 7)
            basis = cls._cell(ws, row, 8)

            if not role_name and not duty and not bottom_line:
                continue

            lines = [
                "## 标准检索单元",
                f"- 数据类型: 岗位义务",
                f"- 部门(sheet): {sheet_name}",
                f"- 部门(行内): {dept_name}",
                f"- 岗位名称: {role_name}",
                f"- 风险内控合规职责: {duty}",
                f"- 合规义务来源(外部): {source_external}",
                f"- 合规义务来源(内部): {source_internal}",
                f"- 合规底线清单: {bottom_line}",
                f"- 底线标准与处罚: {penalty}",
                f"- 制度依据: {basis}",
                "---",
                "",
            ]
            records.append("\n".join(lines))
        return records

    @classmethod
    def _extract_process_records(cls, ws, sheet_name: str) -> list[str]:
        records: list[str] = []
        p1 = ""
        p2 = ""
        p3 = ""
        for row in range(4, ws.max_row + 1):
            p1 = cls._cell(ws, row, 1) or p1
            p2 = cls._cell(ws, row, 2) or p2
            p3 = cls._cell(ws, row, 3) or p3
            process_code = cls._cell(ws, row, 4)
            process_name = cls._cell(ws, row, 5)
            risk_name = cls._cell(ws, row, 6)
            key_point = cls._cell(ws, row, 7)
            source_external = cls._cell(ws, row, 8)
            source_internal = cls._cell(ws, row, 9)
            risk_point = cls._cell(ws, row, 10)

            if not process_code:
                continue

            lines = [
                "## 标准检索单元",
                f"- 数据类型: 流程管控",
                f"- 部门: {sheet_name}",
                f"- 一级流程: {p1}",
                f"- 二级流程: {p2}",
                f"- 三级流程: {p3}",
                f"- 末级流程编号: {process_code}",
                f"- 末级流程名称: {process_name}",
                f"- 合规重要环节(主要风险名称): {risk_name}",
                f"- 合规审查内容(关键控制点): {key_point}",
                f"- 合规义务来源(外部): {source_external}",
                f"- 合规义务来源(内部): {source_internal}",
                f"- 合规风险点: {risk_point}",
                "---",
                "",
            ]
            records.append("\n".join(lines))
        return records

    @classmethod
    async def _import_generated_files(
        cls,
        db_id: str,
        files: list[Path],
        operator_id: str,
        chunk_params: dict,
        force: bool,
    ) -> tuple[int, int]:
        if not files:
            return 0, 0

        db_info = await knowledge_base.get_database_info(db_id)
        file_map = {
            str(meta.get("filename") or "").lower(): (file_id, meta)
            for file_id, meta in (db_info.get("files") or {}).items()
            if meta.get("filename")
        }

        imported = 0
        skipped = 0
        for file_path in files:
            filename = file_path.name.lower()
            existing = file_map.get(filename)

            if existing:
                file_id, meta = existing
                status = meta.get("status")
                if status == FileStatus.INDEXED and not force:
                    skipped += 1
                    continue
                try:
                    if status in {FileStatus.UPLOADED, FileStatus.ERROR_PARSING, FileStatus.PARSING, "failed"}:
                        await knowledge_base.parse_file(db_id, file_id, operator_id=operator_id)
                    await knowledge_base.update_file_params(db_id, file_id, chunk_params, operator_id=operator_id)
                    await knowledge_base.index_file(db_id, file_id, operator_id=operator_id)
                    imported += 1
                except Exception as exc:  # noqa: BLE001
                    logger.error("Seed reprocess failed db_id={} file={} error={}", db_id, file_id, exc)
                continue

            try:
                file_meta = await knowledge_base.add_file_record(
                    db_id=db_id,
                    item=str(file_path),
                    params=chunk_params,
                    operator_id=operator_id,
                )
                file_id = file_meta["file_id"]
                await knowledge_base.parse_file(db_id, file_id, operator_id=operator_id)
                await knowledge_base.update_file_params(db_id, file_id, chunk_params, operator_id=operator_id)
                await knowledge_base.index_file(db_id, file_id, operator_id=operator_id)
                imported += 1
            except Exception as exc:  # noqa: BLE001
                logger.error("Seed import failed db_id={} file={} error={}", db_id, file_path, exc)

        return imported, skipped

    @classmethod
    def sync_source_files(cls, source_dir: Path) -> tuple[Path, list[Path]]:
        """将源excel同步到项目种子目录，便于容器启动时自动导入。"""
        target_dir = Path(config.save_dir) / "jingzhou_compliance_seed" / "source"
        target_dir.mkdir(parents=True, exist_ok=True)
        copied: list[Path] = []
        for kb_def in cls.KB_DEFS:
            src = source_dir / kb_def["source_file"]
            if not src.exists():
                continue
            dst = target_dir / kb_def["source_file"]
            shutil.copy2(src, dst)
            copied.append(dst)
        return target_dir, copied
